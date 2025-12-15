import os, sys
import logging
import re

import polars as pl  # pip install polars
import pandas as pd 
import numpy as np

from pathlib import Path

from rich.console import Console  # pip install rich 
from rich.panel import Panel
from tqdm import tqdm
from tabulate import tabulate

# Create an instance of Console
console = Console()


def print_with_indent_v0(message, indent_level=0, indent_size=4):
    indent = ' ' * (indent_level * indent_size)
    console.print(f"{indent}{message}")


def print_with_indent(message, indent_level=0, indent_size=4):
    """
    Print a message with a specified indentation level using rich.console.

    Parameters:
    - message (str): The message to print.
    - indent_level (int): The level of indentation (default is 0).
    - indent_size (int): The number of spaces per indentation level (default is 4).
    """
    indent = ' ' * (indent_level * indent_size)
    for line in message.split('\n'):
        console.print(f"{indent}{line}")


def print_section_separator(light=False):
    """
    Print a section separator.

    Parameters:
    - light (bool): If True, print a lighter version of the separator.
    """
    if light:
        separator = "." * 85
    else:
        separator = "-" * 85
    console.print(Panel(separator, style="bold"))


def print_emphasized(text, style='bold', edge_effect=True, symbol='='):
    styles = {
        'bold': '\033[1m',
        'underline': '\033[4m',
        'red': '\033[91m',
        'green': '\033[92m',
        'yellow': '\033[93m',
        'blue': '\033[94m',
        'magenta': '\033[95m',
        'cyan': '\033[96m'
    }
    end_style = '\033[0m'  # Reset to default style

    if edge_effect:
        edge_line = symbol * len(text)
        print(edge_line)
        print(f"{styles.get(style, '')}{text}{end_style}")
        print(edge_line)
    else:
        print(f"{styles.get(style, '')}{text}{end_style}")


# Function to display DataFrame in a nice format
def display_dataframe(df, title=None):
    if title:
        print_emphasized(title)
    if isinstance(df, pl.DataFrame):
        df = df.to_pandas()
    print(tabulate(df, headers='keys', tablefmt='psql'))


def display(df, title=None): 
    display_dataframe(df, title)


def display_dataframe_in_chunks(df, num_rows=5, num_columns=6, title=None):
    """
    Display example rows from the DataFrame, including the header, with high readability.

    Parameters:
    - df (pl.DataFrame or pd.DataFrame): The DataFrame to display.
    - num_rows (int): Number of rows to display (default is 5).
    - num_columns (int): Number of columns to display at a time (default is 10).

    Returns:
    - None
    """
    # Convert Polars DataFrame to Pandas DataFrame if necessary
    if isinstance(df, pl.DataFrame):
        df = df.to_pandas()

    if title:
        print_emphasized(title)

    # Display the DataFrame in chunks of columns
    total_columns = df.shape[1]
    for start in range(0, total_columns, num_columns):
        end = start + num_columns
        subset_df = df.iloc[:, start:end]
        print(tabulate(subset_df.head(num_rows), headers=subset_df.columns, tablefmt='psql'))  # headers='keys'
        print("\n")


def setup_logger(log_file_path):
    """
    Set up a logger that prints to the console and writes to a log file.

    Parameters:
    - log_file_path (str): The file path to log warnings.

    """
    logger = logging.getLogger('gene_id_logger')
    logger.setLevel(logging.DEBUG)

    # Create handlers
    console_handler = logging.StreamHandler()
    file_handler = logging.FileHandler(log_file_path)

    # Create formatters and add them to the handlers
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    console_handler.setFormatter(formatter)
    file_handler.setFormatter(formatter)

    # Add handlers to the logger
    logger.addHandler(console_handler)
    logger.addHandler(file_handler)

    return logger


def test_unique_gene_ids(subset_annot_df, subset_seq_df, logger):
    """
    Test for unique gene IDs in annotation and sequence dataframes and log warnings if there is a mismatch.

    Parameters:
    - subset_annot_df (pl.DataFrame): The annotation dataframe.
    - subset_seq_df (pl.DataFrame): The sequence dataframe.
    - logger (logging.Logger): The logger to use for logging warnings.

    Example usage: 
        subset_annot_df = pl.DataFrame({'gene_id': ['gene1', 'gene2', 'gene3']})
        subset_seq_df = pl.DataFrame({'gene_id': ['gene1', 'gene2', 'gene4']})
        logger = setup_logger('gene_id_mismatch.log')
        test_unique_gene_ids(subset_annot_df, subset_seq_df, logger)
    """
    unique_gids_in_annot = set(subset_annot_df.select('gene_id').to_series().to_list())
    unique_gids_in_seq = set(subset_seq_df.select('gene_id').to_series().to_list())

    # Find gene IDs in subset_df but not in annot_df
    missing_gene_ids = set(unique_gids_in_seq) - set(unique_gids_in_annot)

    print_emphasized(f"Number of unique genes in sequence subset: {len(unique_gids_in_seq)}")
    print_with_indent(f"Number of unique genes in annotation subset: {len(unique_gids_in_annot)}", indent_level=1)
    print_with_indent(f"Number of missing gene IDs (in seq but not in annot): {len(missing_gene_ids)}", indent_level=1)

    if unique_gids_in_annot != unique_gids_in_seq:
        warning_message = "[warning] Mismatch in gene IDs between annotation and sequence data"
        difference_message = f"Difference: {unique_gids_in_annot.symmetric_difference(unique_gids_in_seq)}"

        print_emphasized(warning_message)
        print_with_indent(difference_message, indent_level=1)

        # Log the warning messages
        logger.warning(warning_message)
        logger.warning(f"Number of unique genes in sequence subset: {len(unique_gids_in_seq)}")
        logger.warning(f"Number of unique genes in annotation subset: {len(unique_gids_in_annot)}")
        logger.warning(f"Number of missing gene IDs (in seq but not in annot): {len(missing_gene_ids)}")
        logger.warning(difference_message)

    return 


def convert_to_excel_v0(source_dir, target_dir, shorten_sheet_name=False):
    """
    Convert .tsv and .csv files in source_dir to Excel files, organized by chromosome,
    and save outputs in target_dir. Optionally shortens sheet names for readability.

    Parameters:
    - source_dir (str): Directory containing .tsv and .csv files to convert.
    - target_dir (str): Directory to save Excel files.
    - shorten_sheet_name (bool): If True, shorten sheet names to chromosome and partition number.
    """
    import polars as pl
    from glob import glob
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    os.makedirs(target_dir, exist_ok=True)

    # Identify files in source_dir
    performance_files = glob(os.path.join(source_dir, "splice_performance_*.tsv"))
    error_files = glob(os.path.join(source_dir, "splice_errors_*.tsv"))
    summary_files = [
        os.path.join(source_dir, "chromosome_performance_summary.tsv"),
        os.path.join(source_dir, "full_splice_performance.tsv"),
        os.path.join(source_dir, "full_splice_errors.tsv")
    ]

    # Helper function to create shorter names
    def shorten_name(filename):
        match = re.match(r'(splice_(errors|performance))_([0-9XYMT]+)_chunk_(\d+)_\d+', filename)
        if match:
            base_name = match.group(1)  # "splice_errors" or "splice_performance"
            chromosome = match.group(3)  # e.g., "17", "X", "Y", "MT"
            chunk_num = match.group(4)   # chunk number (e.g., "2")
            return f"{base_name}_{chromosome}_{chunk_num}"
        return filename  # Return original name if no match

    # Process performance and error files by chromosome
    chromosomes = {f.split("_")[2] for f in performance_files + error_files}  # Extract unique chromosome identifiers
    for chrom in chromosomes:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])  # Remove default empty sheet only once

        chrom_performance_files = [f for f in performance_files if f"_{chrom}_" in f]
        chrom_error_files = [f for f in error_files if f"_{chrom}_" in f]

        # Add performance files as sheets
        for perf_file in chrom_performance_files:
            sheet_name = shorten_name(os.path.splitext(os.path.basename(perf_file))[0]) if shorten_sheet_name \
                else os.path.splitext(os.path.basename(perf_file))[0]
            df = pl.read_csv(perf_file).to_pandas()
            ws = wb.create_sheet(title=sheet_name)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

        # Add error analysis files as sheets
        for error_file in chrom_error_files:
            sheet_name = shorten_name(os.path.splitext(os.path.basename(error_file))[0]) if shorten_sheet_name \
                else os.path.splitext(os.path.basename(error_file))[0]
            df = pl.read_csv(error_file).to_pandas()
            ws = wb.create_sheet(title=sheet_name)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

        # Save chromosome-specific workbook
        chrom_output_path = os.path.join(target_dir, f"splice_data_chromosome_{chrom}.xlsx")
        wb.save(chrom_output_path)

    # Process the summary files separately
    for summary_file in summary_files:
        file_name = os.path.splitext(os.path.basename(summary_file))[0]
        sheet_name = shorten_name(file_name) if shorten_sheet_name else file_name
        df = pl.read_csv(summary_file).to_pandas()
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        summary_output_path = os.path.join(target_dir, f"{file_name}.xlsx")
        wb.save(summary_output_path)


def detect_delimiter(file_path):
    """
    Detect the delimiter of a file based on its extension or content.
    - If the file has a .tsv extension, assume tab delimiter.
    - If the file has a .csv extension, assume comma delimiter.
    - Otherwise, inspect the first line to determine the delimiter.
    """
    if file_path.endswith(".tsv"):
        return '\t'
    elif file_path.endswith(".csv"):
        return ','
    else:
        # Inspect the first line for non-standard delimiters
        with open(file_path, 'r') as file:
            first_line = file.readline()
            if '\t' in first_line:
                return '\t'
            elif ',' in first_line:
                return ','
            else:
                raise ValueError("Unknown delimiter in file: " + file_path)


# Helper function to create shorter names (v0)
def shorten_name(filename):
    match = re.match(r'(splice_(errors|performance))_([0-9XYMT]+)_chunk_(\d+)_\d+', filename)
    if match:
        base_name = match.group(1)  # "splice_errors" or "splice_performance"
        chromosome = match.group(3)  # chromosome (e.g., "17", "X", "Y", "MT")
        chunk_num = match.group(4)   # chunk number
        return f"{base_name}_{chromosome}_{chunk_num}"
    return filename  # default to full name if pattern does not match

def convert_to_excel(source_dir, target_dir, shorten_sheet_name=False, verbose=1):
    """
    Convert .tsv and .csv files in source_dir to separate Excel files for each chromosome,
    with separate files for performance and error analysis data. Save outputs in target_dir.

    Parameters:
    - source_dir (str): Directory containing .tsv and .csv files to convert.
    - target_dir (str): Directory to save Excel files.
    - shorten_sheet_name (bool): If True, shorten sheet names for readability.
    - verbose (int): If > 0, print detailed progress messages.
    """
    import polars as pl
    from glob import glob
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from tqdm import tqdm
    os.makedirs(target_dir, exist_ok=True)

    if verbose:
        print(f"Source directory: {source_dir}")
        print(f"Target directory: {target_dir}")

    # Identify performance and error files
    performance_files = glob(os.path.join(source_dir, "splice_performance_*.tsv"))
    error_files = glob(os.path.join(source_dir, "splice_errors_*.tsv"))

    # Helper function to create short sheet names based on file order
    def shorten_name(files, base_name):
        sorted_files = sorted(files, key=lambda x: int(re.search(r'chunk_(\d+)', x).group(1)))
        return {f: f"{base_name}_{i+1}" for i, f in enumerate(sorted_files)}

    # Separate files by chromosome and type
    chromosomes = set(re.search(r"_(\d+|X|Y|MT)_", f).group(1) for f in performance_files + error_files)
    for chrom in tqdm(chromosomes, desc="Processing chromosomes"):
        if verbose:
            print_emphasized(f"Processing chromosome {chrom}")

        # Process performance files for the chromosome
        chrom_performance_files = [f for f in performance_files if f"_{chrom}_" in f]
        if chrom_performance_files:
            if verbose:
                print_with_indent(f"Found {len(chrom_performance_files)} performance files for chromosome {chrom}", indent_level=1)
            shortened_perf_names = shorten_name(chrom_performance_files, f"metric_chr{chrom}")

            wb_performance = Workbook()
            if "Sheet" in wb_performance.sheetnames:
                wb_performance.remove(wb_performance["Sheet"])

            for perf_file in chrom_performance_files:
                delimiter = detect_delimiter(perf_file)

                sheet_name = shortened_perf_names[perf_file] if shorten_sheet_name \
                    else os.path.splitext(os.path.basename(perf_file))[0]

                if verbose:
                    print_with_indent(f"Adding sheet '{sheet_name}' from file '{perf_file}'", indent_level=2)

                df = pl.read_csv(perf_file, separator=delimiter).to_pandas()
                ws = wb_performance.create_sheet(title=sheet_name)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)

            # Save the chromosome-specific performance workbook
            performance_output_path = os.path.join(target_dir, f"splice_performance_chr{chrom}.xlsx")
            wb_performance.save(performance_output_path)
            if verbose:
                print_with_indent(f"Saved performance data to '{performance_output_path}'", indent_level=1)

        # Process error files for the chromosome
        chrom_error_files = [f for f in error_files if f"_{chrom}_" in f]
        if chrom_error_files:
            if verbose:
                print_with_indent(f"Found {len(chrom_error_files)} error files for chromosome {chrom}", indent_level=1)
            shortened_error_names = shorten_name(chrom_error_files, f"error_chr{chrom}")

            wb_errors = Workbook()
            if "Sheet" in wb_errors.sheetnames:
                wb_errors.remove(wb_errors["Sheet"])

            for error_file in chrom_error_files:
                delimiter = detect_delimiter(error_file)
                
                sheet_name = shortened_error_names[error_file] if shorten_sheet_name \
                    else os.path.splitext(os.path.basename(error_file))[0]

                if verbose:
                    print_with_indent(f"Adding sheet '{sheet_name}' from file '{error_file}'", indent_level=2)

                df = pl.read_csv(error_file, separator=delimiter).to_pandas()
                ws = wb_errors.create_sheet(title=sheet_name)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)

            # Save the chromosome-specific error workbook
            error_output_path = os.path.join(target_dir, f"splice_errors_chr{chrom}.xlsx")
            wb_errors.save(error_output_path)
            if verbose:
                print_with_indent(f"Saved error data to '{error_output_path}'", indent_level=1)

    # Process summary files separately
    summary_files = [
        os.path.join(source_dir, "chromosome_performance_summary.tsv"),
        os.path.join(source_dir, "full_splice_performance.tsv"),
        os.path.join(source_dir, "full_splice_errors.tsv")
    ]
    for summary_file in summary_files:
        file_name = os.path.splitext(os.path.basename(summary_file))[0]
        sheet_name = file_name 

        if verbose:
            print_emphasized(f"Processing summary file '{summary_file}'")

        df = pl.read_csv(summary_file).to_pandas()
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        summary_output_path = os.path.join(target_dir, f"{file_name}.xlsx")
        wb.save(summary_output_path)

        if verbose:
            print_with_indent(f"Saved summary data to '{summary_output_path}'", indent_level=1)




